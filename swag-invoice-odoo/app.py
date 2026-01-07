import os
import json
import re
from io import BytesIO
from datetime import datetime

import streamlit as st
import pdfplumber
import pandas as pd
from openai import OpenAI
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
)
from reportlab.lib.units import inch

# ---------- CONFIG ----------

st.set_page_config(layout="wide")

HF_TOKEN = "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"

client = OpenAI(
    base_url="https://router.huggingface.co/v1",
    api_key=HF_TOKEN,
)

# Session state init
if "history" not in st.session_state:
    st.session_state["history"] = []
if "ai_cache" not in st.session_state:
    st.session_state["ai_cache"] = {}
if "uploaded_pdf" not in st.session_state:
    st.session_state["uploaded_pdf"] = None
if "uploaded_excel" not in st.session_state:
    st.session_state["uploaded_excel"] = None
if "conversion_mode" not in st.session_state:
    st.session_state["conversion_mode"] = "pdf_to_excel"
if "generated_pdf" not in st.session_state:
    st.session_state["generated_pdf"] = None
if "excel_df" not in st.session_state:
    st.session_state["excel_df"] = None

# ---------- UI: LOGO + PREMIUM, TIGHT LAYOUT CSS ----------

logo_col1, logo_col2, logo_col3 = st.columns([1, 2, 1])
with logo_col2:
    st.markdown(
        "<div style='text-align:center; margin-top:6px; margin-bottom:4px;'>",
        unsafe_allow_html=True,
    )
    st.image(
        "https://raw.githubusercontent.com/sabeya143111-arch/swag-invoice-odoo/main/swag-invoice-odoo/logo.png",
        use_column_width=False,
        width=260,
    )
    st.markdown("</div>", unsafe_allow_html=True)

st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700;800&display=swap');

    .stApp {
        font-family: 'Plus Jakarta Sans', system-ui, -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
        background: #020617;
        color: #e5e7eb;
    }

    .stApp::before {
        content: "";
        position: fixed;
        inset: 0;
        pointer-events: none;
        background:
            radial-gradient(circle at 0% 0%, rgba(56,189,248,0.16) 0, transparent 55%),
            radial-gradient(circle at 100% 100%, rgba(129,140,248,0.16) 0, transparent 55%);
        opacity: 0.9;
    }

    .block-container {
        padding-top: 1.4rem;
        padding-bottom: 1.4rem;
        max-width: 940px;
        margin: 0 auto;
    }

    .main-title {
        font-size: 2.4rem;
        font-weight: 800;
        letter-spacing: 0.08em;
        background: linear-gradient(120deg, #22c55e, #4ade80, #a855f7, #f97316);
        -webkit-background-clip: text;
        color: transparent;
    }
    .sub-text {
        font-size: 0.95rem;
        color: #e5e7eb;
        opacity: 0.9;
    }

    .glass-card {
        background: radial-gradient(circle at top left, rgba(15,23,42,0.98), rgba(15,23,42,0.9));
        border-radius: 20px;
        padding: 20px 22px;
        border: 1px solid rgba(148, 163, 184, 0.5);
        box-shadow: 0 22px 60px rgba(0, 0, 0, 0.8);
        backdrop-filter: blur(20px);
    }

    .pill-badge {
        display: inline-flex;
        align-items: center;
        gap: 8px;
        font-size: 0.78rem;
        padding: 4px 11px;
        border-radius: 999px;
        border: 1px solid rgba(52, 211, 153, 0.75);
        color: #bbf7d0;
        background: linear-gradient(120deg, rgba(22,163,74,0.45), rgba(22,163,74,0.18));
    }

    .stat-card {
        background: radial-gradient(circle at top left, #020617, #020617 55%, #020617);
        border-radius: 16px;
        padding: 14px 16px;
        border: 1px solid rgba(55, 65, 81, 0.9);
        transition: all 0.22s ease;
    }
    .stat-card:hover {
        border-color: rgba(129, 230, 217, 0.9);
        box-shadow: 0 16px 32px rgba(15, 118, 110, 0.55);
        transform: translateY(-2px);
    }

    .stat-label {
        font-size: 0.7rem;
        text-transform: uppercase;
        color: #9ca3af;
        letter-spacing: 0.14em;
    }
    .stat-value {
        font-size: 1.6rem;
        font-weight: 700;
        color: #22c55e;
        margin-top: 2px;
    }

    .dataframe-container {
        border-radius: 14px;
        border: 1px solid rgba(148, 163, 184, 0.6);
        overflow: hidden;
        box-shadow: 0 18px 40px rgba(15, 23, 42, 0.95);
    }

    .success-badge {
        background: linear-gradient(120deg, rgba(22,163,74,0.18), rgba(34,197,94,0.14));
        color: #bbf7d0;
        padding: 11px 14px;
        border-radius: 11px;
        border-left: 4px solid #22c55e;
        font-size: 0.9rem;
        margin: 10px 0;
    }
    .warning-badge {
        background: linear-gradient(120deg, rgba(248,171,89,0.22), rgba(248,113,113,0.15));
        color: #fed7aa;
        padding: 9px 13px;
        border-radius: 11px;
        border-left: 4px solid #f97316;
        font-size: 0.88rem;
        margin: 8px 0;
    }

    .footer-note {
        font-size: 0.8rem;
        color: #9ca3af;
        text-align: center;
        margin-top: 16px;
    }

    .stButton > button {
        font-family: 'Plus Jakarta Sans', system-ui, sans-serif;
        border-radius: 999px;
        padding: 0.48rem 1.3rem;
        border: 1px solid rgba(34,197,94,0.85);
        background: linear-gradient(135deg, #22c55e, #16a34a);
        color: #000000;
        font-weight: 600;
        letter-spacing: 0.04em;
        box-shadow: 0 14px 30px rgba(34,197,94,0.5);
        transition: all 0.2s ease;
        font-size: 0.88rem;
    }
    .stButton > button:hover {
        background: linear-gradient(135deg, #4ade80, #22c55e);
        transform: translateY(-1px);
        box-shadow: 0 20px 40px rgba(34,197,94,0.65);
    }

    .stTextInput > div > input, .stNumberInput input {
        font-family: 'Plus Jakarta Sans', system-ui, sans-serif;
        border-radius: 999px;
        border: 1px solid rgba(148,163,184,0.8);
        background: rgba(15,23,42,0.96);
        color: #e5e7eb;
        font-size: 0.86rem;
        padding-left: 14px;
    }
    .stTextInput > div > input:focus, .stNumberInput input:focus {
        border-color: rgba(56,189,248,0.95);
        box-shadow: 0 0 0 1px rgba(56,189,248,0.95);
    }

    .stTabs [data-baseweb="tab-list"] {
        gap: 0.35rem;
    }
    .stTabs [data-baseweb="tab"] {
        font-family: 'Plus Jakarta Sans', system-ui, sans-serif;
        padding: 0.3rem 0.85rem;
        border-radius: 999px;
        background-color: rgba(15,23,42,0.95);
        border: 1px solid rgba(55,65,81,0.85);
        font-size: 0.8rem;
    }
    .stTabs [aria-selected="true"] {
        background: linear-gradient(120deg, #22c55e, #16a34a);
        border-color: rgba(34,197,94,1);
        color: #000000 !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# ---------- EXCEL -> PDF HELPERS ----------


def detect_excel_structure(df: pd.DataFrame):
    structure = {
        "total_rows": len(df),
        "total_columns": len(df.columns),
        "columns": df.columns.tolist(),
        "has_product": any(
            col.lower()
            in ["product", "model", "product_id", "item", "description"]
            for col in df.columns
        ),
        "has_qty": any(
            col.lower() in ["qty", "quantity", "qua", "units"] for col in df.columns
        ),
        "has_price": any(
            col.lower()
            in ["price", "amount", "cost", "unit_price", "price_unit"]
            for col in df.columns
        ),
    }
    return structure


def excel_to_pdf_with_styling(df: pd.DataFrame, vendor_name: str):
    buffer = BytesIO()

    pdf = SimpleDocTemplate(
        buffer, pagesize=A4, topMargin=0.5 * inch, bottomMargin=0.5 * inch
    )
    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=24,
        textColor=colors.HexColor("#22c55e"),
        spaceAfter=12,
        alignment=1,
        fontName="Helvetica-Bold",
    )

    heading_style = ParagraphStyle(
        "CustomHeading",
        parent=styles["Heading2"],
        fontSize=12,
        textColor=colors.HexColor("#22c55e"),
        spaceAfter=8,
        spaceBefore=8,
        fontName="Helvetica-Bold",
    )

    normal_style = ParagraphStyle(
        "CustomNormal",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.HexColor("#e5e7eb"),
    )

    story.append(Paragraph("📄 INVOICE DOCUMENT", title_style))
    story.append(Spacer(1, 0.2 * inch))

    header_data = [
        ["VENDOR / SUPPLIER", vendor_name],
        ["GENERATED ON", datetime.now().strftime("%d-%b-%Y %H:%M")],
        ["TOTAL ROWS", str(len(df))],
    ]

    header_table = Table(header_data, colWidths=[2 * inch, 3 * inch])
    header_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#16a34a")),
                ("TEXTCOLOR", (0, 0), (0, -1), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (0, -1), 9),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#374151")),
            ]
        )
    )
    story.append(header_table)
    story.append(Spacer(1, 0.3 * inch))

    story.append(Paragraph("📋 LINE ITEMS", heading_style))

    table_data = [df.columns.tolist()]
    for _, row in df.iterrows():
        table_data.append([str(val)[:30] for val in row.values])

    num_cols = len(df.columns)
    if num_cols == 0:
        num_cols = 1
        table_data = [["NO DATA"]]
    col_widths = [7.5 * inch / num_cols] * num_cols

    data_table = Table(table_data, colWidths=col_widths)
    data_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#22c55e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                ("ALIGN", (0, 1), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.HexColor("#0f172a"), colors.HexColor("#1e293b")],
                ),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#e5e7eb")),
                ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#374151")),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(data_table)
    story.append(Spacer(1, 0.3 * inch))

    footer_text = (
        f"<i>Auto-generated PDF from Excel | {datetime.now().strftime('%Y-%m-%d')}</i>"
    )
    story.append(Paragraph(footer_text, normal_style))

    pdf.build(story)
    buffer.seek(0)
    return buffer


def style_excel_file(buffer):
    from openpyxl import load_workbook

    wb = load_workbook(buffer)
    ws = wb.active

    header_fill = PatternFill(
        start_color="22C55E", end_color="22C55E", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    light_fill = PatternFill(
        start_color="F0F9FF", end_color="F0F9FF", fill_type="solid"
    )
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=2, max_row=ws.max_row), start=2
    ):
        if row_idx % 2 == 0:
            for cell in row:
                cell.fill = light_fill
                cell.alignment = Alignment(horizontal="left", vertical="center")

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(buffer)
    buffer.seek(0)

# ---------- PDF PARSE HELPERS ----------


def detect_pdf_structure(text: str):
    lines = [" ".join(ln.split()) for ln in text.split("\n") if ln.strip()]
    structure = {
        "has_sr": any("SR" in ln for ln in lines[:50]),
        "has_model": any(re.search(r"[A-Z]{2,}\-\d+", ln) for ln in lines[:50]),
        "has_qty": any(
            re.search(r"\bQty\b|\bQuantity\b|\bQTY\b", ln, re.I)
            for ln in lines[:50]
        ),
        "has_price": any(
            re.search(r"\bPrice\b|\bAmount\b|\bCost\b", ln, re.I)
            for ln in lines[:50]
        ),
        "total_lines": len(lines),
    }
    return structure, lines


def extract_item_lines_generic(text: str, structure: dict):
    lines = [" ".join(ln.split()) for ln in text.split("\n") if ln.strip()]
    item_lines = []

    if structure["has_sr"]:
        for ln in lines:
            if "SR" in ln:
                sr_amounts = re.findall(r"SR\s*([\d,]+\.\d+)", ln)
                if len(sr_amounts) >= 1 and re.search(
                    r"[A-Za-z0-9\-]+\s+\d+$", ln
                ):
                    item_lines.append(("sr_format", ln))
    else:
        for ln in lines:
            if re.search(r"\d+", ln) and any(c.isalpha() for c in ln):
                if not re.search(r"(total|subtotal|invoice|date)", ln, re.I):
                    item_lines.append(("generic_format", ln))

    return item_lines


def parse_line_sr_format(ln: str):
    sr_amounts = re.findall(r"SR\s*([\d,]+\.\d+)", ln)
    unit_price = float(sr_amounts[-1].replace(",", "")) if sr_amounts else 0.0

    after_last_sr = re.split(r"SR\s*[\d,]+\.\d+", ln)[-1].strip()
    qty_match = re.search(r"(\d+)", after_last_sr)
    qty = float(qty_match.group(1)) if qty_match else 0.0

    model_line_match = re.search(r"([A-Za-z0-9\-]+)\s+(\d+)$", after_last_sr)
    model = model_line_match.group(1) if model_line_match else ""

    tmp = after_last_sr
    if qty_match:
        tmp = re.sub(rf"^{qty_match.group(1)}\s*", "", tmp)
    if model_line_match:
        tmp = tmp.replace(model_line_match.group(0), "")
    desc = " ".join(tmp.split())

    return model.strip(), desc.strip(), qty, unit_price


def parse_line_generic(ln: str):
    numbers = re.findall(r"[\d,]+\.?\d*", ln)
    model_match = re.search(r"([A-Z]{2,}\-?\d+)", ln)
    model = model_match.group(1) if model_match else ""

    desc = re.sub(r"[A-Z]{2,}\-?\d+", "", ln).strip()
    desc = re.sub(r"[\d,]+\.?\d*", "", desc).strip()
    desc = " ".join(desc.split())

    qty = float(numbers[-2].replace(",", "")) if len(numbers) >= 2 else 0.0
    unit_price = float(numbers[-1].replace(",", "")) if len(numbers) >= 1 else 0.0

    return model, desc, qty, unit_price


def pdf_to_odoo_df(
    pdf_file, vendor_name="SWAG TRADING CO.", discount_pct=0.0, vat_pct=0.0
):
    with pdfplumber.open(pdf_file) as pdf:
        full_text = ""
        for page in pdf.pages:
            full_text += (page.extract_text() or "") + "\n"

    structure, lines = detect_pdf_structure(full_text)
    item_lines = extract_item_lines_generic(full_text, structure)

    records = []
    discount_factor = 1 - (discount_pct / 100)
    vat_factor = 1 + (vat_pct / 100)

    for fmt, ln in item_lines:
        if fmt == "sr_format":
            model, desc, qty, price = parse_line_sr_format(ln)
        else:
            model, desc, qty, price = parse_line_generic(ln)

        if not model:
            continue

        line_base = qty * price
        line_total = line_base * discount_factor * vat_factor
        records.append(
            {
                "partner_id/name": vendor_name,
                "order_line/product_id": model,
                "order_line/name": desc,
                "order_line/product_uom_qty": qty,
                "order_line/price_unit": price,
                "order_line/price_subtotal": line_total,
            }
        )

    df = pd.DataFrame(records)
    return df, full_text, item_lines, structure


def analyze_invoice_with_ai(
    df: pd.DataFrame, vendor: str, discount: float, vat: float
) -> str:
    if not HF_TOKEN:
        return "❌ HF_TOKEN missing. Streamlit secrets me set karo."

    sample = df.head(60).to_dict(orient="records")

    prompt = f"""
You are a purchase & inventory analyst for a fashion retail company in Saudi Arabia.

Vendor: {vendor}
Global discount: {discount} %
VAT: {vat} %

Below is invoice line data as JSON list. Each line has:
product_id, description, quantity, unit_price, subtotal.

DATA:
{json.dumps(sample, ensure_ascii=False)}

Tasks (answer in short bullet points, simple English + little Hindi/Urdu mix):
1) Overall summary: total items, total quantity, total amount (approx).
2) Top 5 high-value models with qty & amount.
3) Any suspicious points: very high qty, strange price, duplicates, etc.
4) Suggestions: stock planning / reorder / price check ideas.

Keep answer concise, max 20 bullets.
"""

    try:
        completion = client.chat.completions.create(
            model="moonshotai/Kimi-K2-Instruct-0905",
            messages=[{"role": "user", "content": prompt}],
        )
        return completion.choices[0].message.content
    except Exception as e:
        return f"❌ AI request failed: {e}"

# ---------- MODE SWITCH UI ----------

st.markdown("### 🔄 Select Conversion Mode")

col_mode1, col_mode2 = st.columns(2)
with col_mode1:
    if st.button("📄 → 📊 PDF to Excel", key="pdf_mode", use_container_width=True):
        st.session_state["conversion_mode"] = "pdf_to_excel"
        st.rerun()

with col_mode2:
    if st.button("📊 → 📄 Excel to PDF", key="excel_mode", use_container_width=True):
        st.session_state["conversion_mode"] = "excel_to_pdf"
        st.rerun()

st.markdown("---")

# ========== MODE 1: PDF TO EXCEL ==========

if st.session_state["conversion_mode"] == "pdf_to_excel":
    left, right = st.columns([1.25, 1])

    with left:
        st.markdown(
            """
            <div class="glass-card">
                <div class="pill-badge">
                    🧾 Auto Invoice → Odoo
                    <span style="opacity:0.7;">• SWAG internal tool</span>
                </div>
                <div style="margin-top: 6px;"></div>
                <div class="main-title">
                    SWAG Invoice → Odoo Excel + AI
                </div>
                <p class="sub-text">
                    Kisi bhi PDF invoice upload karo, app automatically structure detect karke 
                    clean Excel bana dega jo direct Odoo import me use ho sakta hai.
                    Upar se AI tumhare data ke hisaab se summary, issues aur suggestions bhi dega.
                </p>
            </div>
            """,
            unsafe_allow_html=True,
        )

        st.write("")
        vendor_name = st.text_input(
            "Vendor / Partner name",
            value="SWAG TRADING CO.",
            help="Odoo ka vendor / partner name likho.",
        )

        uploaded_pdf_widget = st.file_uploader(
            "Invoice PDF upload karein (kisi bhi format)",
            type=["pdf"],
            help="SWAG ya kisi aur supplier ka invoice (PDF).",
        )

        if uploaded_pdf_widget is not None:
            st.session_state["uploaded_pdf"] = uploaded_pdf_widget

        uploaded_pdf = st.session_state["uploaded_pdf"]

        with st.expander("⚙️ Advanced settings", expanded=False):
            discount_pct = st.number_input(
                "Global discount %", 0.0, 100.0, 0.0, 0.5
            )
            vat_pct = st.number_input("VAT %", 0.0, 30.0, 0.0, 0.5)

        convert_clicked = st.button("🔁 Convert to Odoo Excel")

    with right:
        st.markdown(
            """
            <div class="glass-card">
                <div class="stat-card">
                    <div class="stat-label">Status</div>
            """,
            unsafe_allow_html=True,
        )

        if uploaded_pdf is None:
            st.markdown(
                """
                    <div class="stat-value">No file uploaded</div>
                    <div style="font-size:0.8rem;color:#9ca3af;margin-top:4px;">
                        Pehle left side se invoice PDF choose karein.
                    </div>
                </div>
            """,
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                f"""
                    <div class="stat-value">Ready to convert</div>
                    <div style="font-size:0.8rem;color:#9ca3af;margin-top:4px;">
                        File: <span style="color:#22c55e;">{uploaded_pdf.name}</span>
                    </div>
                </div>
            """,
                unsafe_allow_html=True,
            )

        st.markdown("</div>", unsafe_allow_html=True)

    tab_overview, tab_details, tab_ai, tab_debug = st.tabs(
        ["📊 Overview", "📋 Details", "🤖 AI Insights", "🛠 Debug"]
    )

    df_odoo = None
    full_text = ""
    item_lines = []
    detected_structure = None

    if uploaded_pdf is not None and convert_clicked:
        progress = st.progress(0, text="Step 1/3: PDF read ho raha hai...")

        with st.spinner("📄 PDF parse aur structure detect ho rahi hai..."):
            progress.progress(
                30, text="Step 2/3: Structure detect + data extract..."
            )
            df_odoo, full_text, item_lines, detected_structure = pdf_to_odoo_df(
                uploaded_pdf, vendor_name, discount_pct, vat_pct
            )
            progress.progress(70, text="Step 3/3: Excel build ho raha hai...")

        if df_odoo is None or df_odoo.empty:
            progress.empty()
            st.error(
                "❌ Koi item line detect nahi hui. PDF format ya structure check karein."
            )
        else:
            total_items = len(df_odoo)
            total_qty = float(df_odoo["order_line/product_uom_qty"].sum())
            total_subtotal = float(df_odoo["order_line/price_subtotal"].sum())
            total_unit_sum = float(df_odoo["order_line/price_unit"].sum())

            progress.progress(100, text="✅ Ho gaya!")
            progress.empty()

            st.session_state["history"].insert(
                0,
                {
                    "File": uploaded_pdf.name,
                    "Items": int(total_items),
                    "Qty": int(total_qty),
                    "Amount": round(total_subtotal, 2),
                },
            )
            st.session_state["history"] = st.session_state["history"][:5]

            with tab_overview:
                format_detected = (
                    "SR-based format (SWAG original)"
                    if detected_structure.get("has_sr")
                    else "Generic flexible format"
                )
                st.markdown(
                    f"""
                    <div class="success-badge">
                        ✅ <strong>{total_items} items successfully extracted</strong> ({format_detected})
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

                dupes = df_odoo["order_line/product_id"].value_counts()
                dupe_models = dupes[dupes > 1]
                if not dupe_models.empty:
                    ex_models = ", ".join(dupe_models.index.tolist()[:3])
                    st.markdown(
                        f"""
                        <div class="warning-badge">
                            ⚠️ {len(dupe_models)} models repeated in invoice. 
                            Example: {ex_models}
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )

                c1, c2, c3, c4 = st.columns(4)
                with c1:
                    st.markdown(
                        f"""
                        <div class="stat-card">
                            <div class="stat-label">📦 Total Items</div>
                            <div class="stat-value">{total_items}</div>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )
                with c2:
                    st.markdown(
                        f"""
                        <div class="stat-card">
                            <div class="stat-label">📊 Total Quantity</div>
                            <div class="stat-value">{total_qty:.0f}</div>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )
                with c3:
                    st.markdown(
                        f"""
                        <div class="stat-card">
                            <div class="stat-label">💰 Unit Price Sum</div>
                            <div class="stat-value">SR {total_unit_sum:,.0f}</div>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )
                with c4:
                    st.markdown(
                        f"""
                        <div class="stat-card">
                            <div class="stat-label">✨ Total Amount</div>
                            <div class="stat-value">SR {total_subtotal:,.0f}</div>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )

                st.markdown(
                    f"""
                    <p style="font-size:0.9rem;color:#9ca3af;margin-top:6px;">
                        Vendor <b>{vendor_name}</b> • Discount <b>{discount_pct:.1f}%</b> • VAT <b>{vat_pct:.1f}%</b>
                    </p>
                    """,
                    unsafe_allow_html=True,
                )

                if st.session_state["history"]:
                    st.markdown("### 🕒 Recent conversions")
                    st.table(st.session_state["history"])

            with tab_details:
                st.markdown("### 🔍 Filters")
                f1, f2 = st.columns(2)
                with f1:
                    min_qty = st.number_input(
                        "Minimum quantity",
                        min_value=0.0,
                        value=0.0,
                        step=1.0,
                        key="min_qty",
                    )
                with f2:
                    min_amount = st.number_input(
                        "Minimum line amount (SR)",
                        min_value=0.0,
                        value=0.0,
                        step=10.0,
                        key="min_amt",
                    )

                filtered_df = df_odoo[
                    (df_odoo["order_line/product_uom_qty"] >= min_qty)
                    & (df_odoo["order_line/price_subtotal"] >= min_amount)
                ]

                st.markdown("### 📋 Preview (Filtered lines)")
                st.markdown(
                    '<div class="dataframe-container">', unsafe_allow_html=True
                )
                st.dataframe(filtered_df, use_container_width=True)
                st.markdown("</div>", unsafe_allow_html=True)

                top5 = df_odoo.sort_values(
                    "order_line/price_subtotal", ascending=False
                ).head(5)
                st.markdown("#### 🔝 Top 5 high value lines")
                st.dataframe(top5, use_container_width=True, height=250)

                buffer = BytesIO()
                with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                    df_odoo.to_excel(
                        writer, index=False, sheet_name="Purchase Orders"
                    )
                style_excel_file(buffer)

                st.download_button(
                    label="⬇️ Download Styled Excel (Ready for Odoo Import)",
                    data=buffer,
                    file_name="odoo_purchase_orders.xlsx",
                    mime=(
                        "application/"
                        "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    ),
                )

                st.markdown(
                    """
                    <div class="footer-note">
                        💡 Excel file automatically color-coded aur formatted hai, 
                        bilkul Odoo import ke liye ready!
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

            with tab_ai:
                st.markdown("### 🤖 AI Insights (invoice ke hisaab se)")

                if not HF_TOKEN:
                    st.error(
                        "HF_TOKEN set nahi hai. `.streamlit/secrets.toml` me HF_TOKEN daalo."
                    )
                else:
                    key = f"{uploaded_pdf.name}_{total_items}_{total_subtotal}"
                    generate_clicked = st.button("Generate AI Insights")

                    if generate_clicked:
                        if key in st.session_state["ai_cache"]:
                            ai_text = st.session_state["ai_cache"][key]
                        else:
                            with st.spinner(
                                "AI soch raha hai... (thoda time lag sakta hai)"
                            ):
                                ai_text = analyze_invoice_with_ai(
                                    df_odoo,
                                    vendor_name,
                                    discount_pct,
                                    vat_pct,
                                )
                                st.session_state["ai_cache"][key] = ai_text
                        st.markdown(ai_text)

                    st.markdown(
                        """
                        <div class="footer-note">
                            Note: Ye AI sirf helper hai – final decision hamesha 
                            tumhare business logic ke hisaab se lo.
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )

            with tab_debug:
                st.markdown("### 🔍 Detected PDF Structure")
                st.json(detected_structure)

                st.markdown("#### Raw text (first 2500 chars)")
                st.code(full_text[:2500])

                st.markdown("#### Detected item lines")
                st.write(f"Total lines detected: {len(item_lines)}")
                for fmt, ln in item_lines[:10]:
                    st.caption(f"[{fmt}] {ln}")

    elif uploaded_pdf is None:
        with tab_overview:
            st.info("📂 Upar se PDF select karo start karne ke liye.")

# ========== MODE 2: EXCEL TO PDF ==========

else:
    left, right = st.columns([1.25, 1])

    with left:
        st.markdown(
            """
            <div class="glass-card">
                <div class="pill-badge">
                    📊 Excel → PDF
                    <span style="opacity:0.7;">• Auto-format invoice</span>
                </div>
                <div style="margin-top: 6px;"></div>
                <div class="main-title">
                    Excel to PDF Invoice
                </div>
                <p class="sub-text">
                    Kisi bhi Excel file (kisi bhi structure mein) upload karo.
                    App automatically detect karega aur professional PDF invoice bana dega.
                    Data kisi bhi format mein ho sakta hai – app smart format karega!
                </p>
            </div>
            """,
            unsafe_allow_html=True,
        )

        st.write("")
        vendor_name_excel = st.text_input(
            "Vendor / Company name",
            value="SWAG TRADING CO.",
            help="Invoice mein vendor ka name dikhega",
        )

        uploaded_excel_widget = st.file_uploader(
            "Excel file upload karein (XLS, XLSX)",
            type=["xlsx", "xls"],
            help="Kisi bhi structure ka Excel file - app smart detect karega",
            key="excel_uploader",
        )

        if uploaded_excel_widget is not None:
            st.session_state["uploaded_excel"] = uploaded_excel_widget

        uploaded_excel = st.session_state["uploaded_excel"]

        convert_excel_clicked = st.button(
            "🔁 Convert to PDF Invoice", key="excel_convert"
        )

    with right:
        st.markdown(
            """
            <div class="glass-card">
                <div class="stat-card">
                    <div class="stat-label">Status</div>
            """,
            unsafe_allow_html=True,
        )

        if uploaded_excel is None:
            st.markdown(
                """
                    <div class="stat-value">No file uploaded</div>
                    <div style="font-size:0.8rem;color:#9ca3af;margin-top:4px;">
                        Pehle left side se Excel choose karein.
                    </div>
                </div>
            """,
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                f"""
                    <div class="stat-value">Ready to convert</div>
                    <div style="font-size:0.8rem;color:#9ca3af;margin-top:4px;">
                        File: <span style="color:#22c55e;">{uploaded_excel.name}</span>
                    </div>
                </div>
            """,
                unsafe_allow_html=True,
            )

        st.markdown("</div>", unsafe_allow_html=True)

    tab_excel_overview, tab_excel_preview, tab_excel_download = st.tabs(
        ["📊 Overview", "👁️ Preview", "⬇️ Download"]
    )

    if uploaded_excel is not None and convert_excel_clicked:
        progress = st.progress(0, text="Excel read ho raha hai...")

        try:
            df_excel = pd.read_excel(uploaded_excel)
            progress.progress(30, text="Structure detect ho raha hai...")

            excel_structure = detect_excel_structure(df_excel)
            progress.progress(60, text="PDF generate ho raha hai...")

            pdf_buffer = excel_to_pdf_with_styling(df_excel, vendor_name_excel)
            progress.progress(100, text="✅ PDF ready!")
            progress.empty()

            st.session_state["generated_pdf"] = pdf_buffer
            st.session_state["excel_df"] = df_excel

            with tab_excel_overview:
                st.markdown(
                    """
                    <div class="success-badge">
                        ✅ <strong>PDF successfully generated!</strong>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

                c1, c2, c3 = st.columns(3)
                with c1:
                    st.markdown(
                        f"""
                        <div class="stat-card">
                            <div class="stat-label">📋 Total Rows</div>
                            <div class="stat-value">{excel_structure['total_rows']}</div>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )
                with c2:
                    st.markdown(
                        f"""
                        <div class="stat-card">
                            <div class="stat-label">📊 Columns</div>
                            <div class="stat-value">{excel_structure['total_columns']}</div>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )
                with c3:
                    st.markdown(
                        """
                        <div class="stat-card">
                            <div class="stat-label">📄 Format</div>
                            <div class="stat-value">Professional</div>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )

                st.markdown("### 📝 Detected Structure")
                struct_col1, struct_col2 = st.columns(2)
                with struct_col1:
                    st.write("**Columns Found:**")
                    for col in excel_structure["columns"]:
                        st.caption(f"• {col}")
                with struct_col2:
                    st.write("**Features Detected:**")
                    st.caption(
                        f"✓ Has Product: {excel_structure['has_product']}"
                    )
                    st.caption(f"✓ Has Quantity: {excel_structure['has_qty']}")
                    st.caption(f"✓ Has Price: {excel_structure['has_price']}")

            with tab_excel_preview:
                st.markdown("### 👁️ Data Preview")
                st.dataframe(
                    st.session_state.get("excel_df", df_excel),
                    use_container_width=True,
                )

            with tab_excel_download:
                st.markdown("### ⬇️ Download Your PDF")
                pdf_buffer_download = st.session_state.get(
                    "generated_pdf", pdf_buffer
                )
                pdf_buffer_download.seek(0)

                st.download_button(
                    label="📥 Download PDF Invoice",
                    data=pdf_buffer_download,
                    file_name=f"invoice_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                    mime="application/pdf",
                )

                st.markdown(
                    """
                    <div class="footer-note">
                        💡 PDF professionally formatted aur print-ready hai!
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

        except Exception as e:
            progress.empty()
            st.error(f"❌ Error: {str(e)}")
            st.info("Excel file properly formatted hai check karo")

    else:
        if uploaded_excel is None:
            with tab_excel_overview:
                st.info("📂 Upar se Excel file select karo start karne ke liye.")
